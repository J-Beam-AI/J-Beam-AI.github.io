import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Colour palette ──────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"   # title bar
MID_BLUE    = "2E75B6"   # section header rows (INCOME / EXPENSES / SUMMARY)
LIGHT_BLUE  = "BDD7EE"   # column-header row
CATEGORY    = "D6E4F0"   # sub-category rows (HOUSING, FOOD, …)
GREEN       = "E2EFDA"   # income data rows
EXPENSE_ROW = "FFF2CC"   # expense data rows
TOTAL_ROW   = "FCE4D6"   # total rows
SUMMARY_ROW = "EAF1FB"   # summary data rows
WHITE       = "FFFFFF"

# ── Border helper ────────────────────────────────────────────────────────────
thin  = Side(style="thin",   color="AAAAAA")
thick = Side(style="medium", color="888888")

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def thick_border():
    return Border(left=thick, right=thick, top=thick, bottom=thick)

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

CURRENCY_FMT = '"$"#,##0.00_);[Red]("$"#,##0.00)'

def apply_currency(cell):
    cell.number_format = CURRENCY_FMT

def style_row(ws, row, cols, bg, fnt, aln=None, border=True):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg)
        cell.font = fnt
        if aln:
            cell.alignment = aln
        if border:
            cell.border = thin_border()

# ────────────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Monthly Budget"

# ── Column widths ────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 28

# ── Row 1 – Title ─────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 34
ws.merge_cells("A1:E1")
t = ws["A1"]
t.value     = "HOME FINANCE TRACKER"
t.font      = Font(name="Calibri", bold=True, size=20, color=WHITE)
t.fill      = fill(DARK_BLUE)
t.alignment = align("center")

# ── Row 2 – Month / Year ──────────────────────────────────────────────────────
ws.row_dimensions[2].height = 20
ws["A2"].value = "Month:"
ws["B2"].value = "January"
ws["C2"].value = "Year:"
ws["D2"].value = 2026
for col in "ABCDE":
    c = ws[f"{col}2"]
    c.fill = fill(LIGHT_BLUE)
    c.font = font(bold=True, size=11)
    c.alignment = align("center")
    c.border = thin_border()

# ── Row 3 – blank ─────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# INCOME SECTION  (rows 4-9)
# ─────────────────────────────────────────────────────────────────────────────
# Row 4 – section header
ws.row_dimensions[4].height = 18
ws.merge_cells("A4:E4")
h = ws["A4"]
h.value     = "INCOME"
h.font      = font(bold=True, size=12, color=WHITE)
h.fill      = fill(MID_BLUE)
h.alignment = align("center")

# Row 5 – column headers
ws.row_dimensions[5].height = 16
for col, label in zip("ABCDE", ["Category", "Budgeted", "Actual", "Difference", "Notes"]):
    c = ws[f"{col}5"]
    c.value     = label
    c.font      = font(bold=True, size=10)
    c.fill      = fill(LIGHT_BLUE)
    c.alignment = align("center")
    c.border    = thin_border()

# Income data rows 6-9
income_items = [
    "Primary Salary",
    "Secondary Income",
    "Freelance / Side Jobs",
    "Other Income",
]
for i, item in enumerate(income_items, start=6):
    ws.cell(row=i, column=1).value = item
    ws.cell(row=i, column=2).value = 0
    ws.cell(row=i, column=3).value = 0
    ws.cell(row=i, column=4).value = f"=C{i}-B{i}"
    style_row(ws, i, range(1, 6), GREEN, font(size=10))
    apply_currency(ws.cell(row=i, column=2))
    apply_currency(ws.cell(row=i, column=3))
    apply_currency(ws.cell(row=i, column=4))
    ws.cell(row=i, column=1).alignment = align("left")
    ws.cell(row=i, column=5).alignment = align("left")

# Row 10 – TOTAL INCOME
ws.row_dimensions[10].height = 18
for col, val in zip(range(1, 6), [
    "TOTAL INCOME",
    "=SUM(B6:B9)",
    "=SUM(C6:C9)",
    "=SUM(D6:D9)",
    ""
]):
    c = ws.cell(row=10, column=col)
    c.value     = val
    c.font      = font(bold=True, size=10, color=WHITE)
    c.fill      = fill(MID_BLUE)
    c.alignment = align("center")
    c.border    = thick_border()
    if col in (2, 3, 4):
        c.number_format = CURRENCY_FMT

# ─────────────────────────────────────────────────────────────────────────────
# EXPENSES SECTION  (rows 12-58)
# ─────────────────────────────────────────────────────────────────────────────
# Row 12 – section header
ws.row_dimensions[12].height = 18
ws.merge_cells("A12:E12")
h = ws["A12"]
h.value     = "EXPENSES"
h.font      = font(bold=True, size=12, color=WHITE)
h.fill      = fill(MID_BLUE)
h.alignment = align("center")

# Row 13 – column headers (repeat)
ws.row_dimensions[13].height = 16
for col, label in zip("ABCDE", ["Category", "Budgeted", "Actual", "Difference", "Notes"]):
    c = ws[f"{col}13"]
    c.value     = label
    c.font      = font(bold=True, size=10)
    c.fill      = fill(LIGHT_BLUE)
    c.alignment = align("center")
    c.border    = thin_border()

# Each sub-section: (category_label, [items])
expense_sections = [
    ("HOUSING", [
        "Rent / Mortgage",
        "Property Tax / HOA",
        "Home Repairs",
    ]),
    ("UTILITIES", [
        "Electricity",
        "Water / Sewer",
        "Gas / Heating",
        "Internet",
        "Phone",
    ]),
    ("FOOD", [
        "Groceries",
        "Dining Out",
    ]),
    ("TRANSPORTATION", [
        "Car Payment",
        "Gas / Fuel",
        "Car Insurance",
        "Public Transit",
        "Car Maintenance",
    ]),
    ("HEALTH", [
        "Health Insurance",
        "Doctor / Dentist",
        "Prescriptions",
        "Gym / Fitness",
    ]),
    ("PERSONAL", [
        "Clothing",
        "Personal Care",
        "Subscriptions",
        "Entertainment",
    ]),
    ("SAVINGS & DEBT", [
        "Emergency Fund",
        "Retirement (401k / IRA)",
        "Credit Card Payment",
        "Student Loans",
    ]),
    ("MISCELLANEOUS", [
        "Gifts / Donations",
        "Education",
        "Pet Expenses",
        "Other",
    ]),
]

current_row = 14
expense_data_rows = []   # track rows with actual data for TOTAL formula

for section_label, items in expense_sections:
    # Category header row
    ws.row_dimensions[current_row].height = 16
    ws.merge_cells(f"A{current_row}:E{current_row}")
    cat = ws.cell(row=current_row, column=1)
    cat.value     = section_label
    cat.font      = font(bold=True, size=10)
    cat.fill      = fill(CATEGORY)
    cat.alignment = align("left")
    cat.border    = thin_border()
    current_row += 1

    # Item rows
    for item in items:
        r = current_row
        expense_data_rows.append(r)
        ws.cell(row=r, column=1).value = item
        ws.cell(row=r, column=2).value = 0
        ws.cell(row=r, column=3).value = 0
        ws.cell(row=r, column=4).value = f"=C{r}-B{r}"
        style_row(ws, r, range(1, 6), EXPENSE_ROW, font(size=10))
        apply_currency(ws.cell(row=r, column=2))
        apply_currency(ws.cell(row=r, column=3))
        apply_currency(ws.cell(row=r, column=4))
        ws.cell(row=r, column=1).alignment = align("left")
        ws.cell(row=r, column=5).alignment = align("left")
        current_row += 1

    current_row += 1  # blank row between sections

# TOTAL EXPENSES row
total_exp_row = current_row
ws.row_dimensions[total_exp_row].height = 18

# Build a sum over only the data rows (avoids accidentally including header rows)
first_exp = expense_data_rows[0]
last_exp  = expense_data_rows[-1]

for col, val in zip(range(1, 6), [
    "TOTAL EXPENSES",
    f"=SUM(B{first_exp}:B{last_exp})",
    f"=SUM(C{first_exp}:C{last_exp})",
    f"=SUM(D{first_exp}:D{last_exp})",
    ""
]):
    c = ws.cell(row=total_exp_row, column=col)
    c.value     = val
    c.font      = font(bold=True, size=10, color=WHITE)
    c.fill      = fill(TOTAL_ROW[:-2] + "00") if col == 5 else fill("C00000")
    c.alignment = align("center")
    c.border    = thick_border()
    if col in (2, 3, 4):
        c.number_format = CURRENCY_FMT

# fix total row to use a deep red background
for col in range(1, 6):
    ws.cell(row=total_exp_row, column=col).fill  = fill("C00000")
    ws.cell(row=total_exp_row, column=col).font  = font(bold=True, size=10, color=WHITE)

# ─────────────────────────────────────────────────────────────────────────────
# SUMMARY SECTION
# ─────────────────────────────────────────────────────────────────────────────
summary_start = total_exp_row + 2   # one blank row gap

ws.row_dimensions[summary_start].height = 18
ws.merge_cells(f"A{summary_start}:E{summary_start}")
s = ws.cell(row=summary_start, column=1)
s.value     = "SUMMARY"
s.font      = font(bold=True, size=12, color=WHITE)
s.fill      = fill(DARK_BLUE)
s.alignment = align("center")

# summary column headers
hdr_row = summary_start + 1
for col, label in zip("ABCDE", ["", "Budgeted", "Actual", "Difference", ""]):
    c = ws[f"{col}{hdr_row}"]
    c.value     = label
    c.font      = font(bold=True, size=10)
    c.fill      = fill(LIGHT_BLUE)
    c.alignment = align("center")
    c.border    = thin_border()

# summary data rows
summary_items = [
    ("Total Income",           f"=B10",             f"=C10",             f"=D10"),
    ("Total Expenses",         f"=B{total_exp_row}", f"=C{total_exp_row}", f"=D{total_exp_row}"),
    ("NET  (Income − Expenses)", None,               None,                None),
]
for offset, (label, bud, act, diff) in enumerate(summary_items):
    r = hdr_row + 1 + offset
    ws.row_dimensions[r].height = 18
    net_row = r  # keep reference for the NET row

    ws.cell(row=r, column=1).value = label
    if bud:
        ws.cell(row=r, column=2).value = bud
        ws.cell(row=r, column=3).value = act
        ws.cell(row=r, column=4).value = diff
    else:
        # NET row – reference the two rows above
        income_r  = hdr_row + 1
        expense_r = hdr_row + 2
        ws.cell(row=r, column=2).value = f"=B{income_r}-B{expense_r}"
        ws.cell(row=r, column=3).value = f"=C{income_r}-C{expense_r}"
        ws.cell(row=r, column=4).value = f"=D{income_r}-D{expense_r}"

    bg = TOTAL_ROW if offset == 2 else SUMMARY_ROW
    style_row(ws, r, range(1, 6), bg, font(bold=(offset == 2), size=10))
    for col in (2, 3, 4):
        apply_currency(ws.cell(row=r, column=col))
    ws.cell(row=r, column=1).alignment = align("left")

# ── Freeze panes below title + header ────────────────────────────────────────
ws.freeze_panes = "A6"

# ── Save ──────────────────────────────────────────────────────────────────────
output_path = r"C:\Users\Jon\Documents\Jon\Artificial Intelligence\Claude Code related\Claude - Vibe Coding\practice\home_finance_tracker.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
